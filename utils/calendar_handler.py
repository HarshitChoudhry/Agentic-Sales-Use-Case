import os
import datetime
import pandas as pd
from typing import List, Dict, Any
from difflib import SequenceMatcher
from google.oauth2.credentials import Credentials
from googleapiclient.discovery import build
from dotenv import load_dotenv

from openpyxl import load_workbook, Workbook

# Load environment variables
load_dotenv()


def normalize(s: str) -> str:
    if s is None:
        return ""
    return str(s).strip().lower().replace("_", " ").replace("-", " ").replace("technologies", "").strip()


def format_name(raw: str) -> str:
    """Format a displayName/email into a neat Title Case name."""
    if not raw:
        return "Unknown"
    if "@" in raw:  # if it's an email
        base = raw.split("@")[0]
    else:
        base = raw
    base = base.replace(".", " ").replace("_", " ")
    return " ".join([w.capitalize() for w in base.split()])


class CalendarHandler:
    """
    Calendar handler that fetches Google Calendar events
    and writes them into meetings.xlsx, cleanly formatting participants.
    """

    def __init__(self, meetings_file: str = "meetings.xlsx", crm_file: str = "crm_data.xlsx"):
        self.meetings_file = meetings_file
        self.crm_file = crm_file
        self.service = None

    def authenticate(self):
        """Authenticate with Google Calendar using .env credentials."""
        if self.service:
            return self.service
        creds = Credentials(
            token=None,
            refresh_token=os.getenv("GOOGLE_REFRESH_TOKEN"),
            token_uri="https://oauth2.googleapis.com/token",
            client_id=os.getenv("GOOGLE_CLIENT_ID"),
            client_secret=os.getenv("GOOGLE_CLIENT_SECRET"),
            scopes=["https://www.googleapis.com/auth/calendar.readonly"],
        )
        self.service = build("calendar", "v3", credentials=creds)
        return self.service

    def fetch_upcoming_events(self, days_ahead: int = 7) -> List[Dict[str, Any]]:
        """Fetch upcoming events within the next N days."""
        svc = self.authenticate()
        now = datetime.datetime.utcnow().isoformat() + "Z"
        future = (datetime.datetime.utcnow() + datetime.timedelta(days=days_ahead)).isoformat() + "Z"

        events_result = (
            svc.events()
            .list(
                calendarId=os.getenv("GOOGLE_CALENDAR_ID", "primary"),
                timeMin=now,
                timeMax=future,
                singleEvents=True,
                orderBy="startTime",
            )
            .execute()
        )

        events = events_result.get("items", [])
        meetings = []

        for event in events:
            title = event.get("summary", "Untitled Meeting")
            start = event["start"].get("dateTime", event["start"].get("date"))
            attendees = []
            if "attendees" in event:
                for att in event["attendees"]:
                    name = att.get("displayName") or att.get("email", "Unknown")
                    attendees.append(name)

            meetings.append({"title": title, "date": start, "participants": attendees})
        return meetings

    def _ensure_workbook(self):
        """Load existing workbook or create new one with sheets for CRM clients."""
        if os.path.exists(self.meetings_file):
            wb = load_workbook(self.meetings_file)
        else:
            wb = Workbook()
            default = wb.active
            wb.remove(default)
            if os.path.exists(self.crm_file):
                crm_df = pd.read_excel(self.crm_file)
                if "Client" in crm_df.columns:
                    for client in crm_df["Client"].dropna().unique():
                        wb.create_sheet(title=str(client))
            if not wb.sheetnames:
                wb.create_sheet(title="Default")
            wb.save(self.meetings_file)
        return wb

    def add_to_meetings_excel(self, events: List[Dict[str, Any]]) -> List[str]:
        """Add fetched events into meetings.xlsx (appends as new rows)."""
        updated = []
        if not events:
            return updated

        # Load CRM for Deal_ID lookup
        crm_df = None
        if os.path.exists(self.crm_file):
            crm_df = pd.read_excel(self.crm_file, dtype=str)
            if "Client" in crm_df.columns:
                crm_df["norm_client"] = crm_df["Client"].astype(str).apply(normalize)

        wb = self._ensure_workbook()
        sheet_names = wb.sheetnames

        for ev in events:
            title = str(ev.get("title", "")).strip()
            date_val = ev.get("date", "")
            participants = ev.get("participants", []) or []

            # --- Find matching client sheet ---
            target_sheet = None
            for s in sheet_names:
                if s.strip().lower() == title.lower():
                    target_sheet = s
                    break
            if not target_sheet:
                norm_title = normalize(title)
                for s in sheet_names:
                    if normalize(s) in norm_title or norm_title in normalize(s):
                        target_sheet = s
                        break
            if not target_sheet:
                print(f"[WARN] No client sheet matched for event '{title}'. Skipping.")
                continue

            ws = wb[target_sheet]

            # Ensure headers
            headers = ["Meeting_ID", "Deal_ID", "Title", "Date", "Participants", "Status", "Transcript_File", "Notes"]
            if ws.max_row == 0:
                ws.append(headers)
            else:
                current_headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
                if not all(h in current_headers for h in headers):
                    ws.delete_rows(1)
                    ws.append(headers)

            # Auto-increment Meeting_ID
            existing_ids = [ws.cell(row=r, column=1).value for r in range(2, ws.max_row + 1)]
            nums = [int(str(mid).replace("M", "")) for mid in existing_ids if mid and str(mid).startswith("M")]
            next_id = f"M{(max(nums) + 1) if nums else 1:03d}"

            # Deal_ID from CRM
            deal_id = ""
            if crm_df is not None:
                match = crm_df[crm_df["norm_client"] == normalize(target_sheet)]
                if not match.empty and "Deal_ID" in match.columns:
                    deal_id = str(match.iloc[0]["Deal_ID"])

            # Format participants (Name + Role)
            formatted = []
            for p in participants:
                name = format_name(p)
                if "aakashgupta" in str(p).lower():
                    formatted.append(f"{name} (Dice)")
                else:
                    formatted.append(f"{name} (Client)")
            participants_str = ", ".join(formatted)

            new_row = [next_id, deal_id, title, date_val, participants_str, "upcoming", "", "Imported from Google Calendar"]
            ws.append(new_row)
            updated.append(target_sheet)

            print(f"✅ Added meeting to {target_sheet}: {new_row}")

        try:
            wb.save(self.meetings_file)
            print(f"[INFO] Saved updates to {self.meetings_file}")
        except PermissionError:
            print(f"[ERROR] Permission denied. Please close {self.meetings_file} in Excel before running again.")

        return updated
